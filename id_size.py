import openpyxl
from openpyxl import Workbook

wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
header_row = []
for i in range(sheet.max_column):
    header_row.append(sheet.cell(row=1, column=i+1).value)
main_l = []
for row in sheet:
    temp_l = []
    for cell in row:
        temp_l.append(cell.value)
    main_l.append(temp_l)
del main_l[0]
wb.close()

print(len(main_l))

wb1 = Workbook()  # Creates a new virtual workbook with 1 sheet by default
ws = wb1.active
clm1, clm2, clm3 = [], [], []
for i in range(len(header_row)-1):
    for row in main_l:
        clm1.append(row[0])
        clm2.append(header_row[i+1])
        clm3.append(row[i+1])

# inserting values to a three columns from our generated three lists
for ind in range(len(clm1)):
    ws.cell(ind + 1, 1, clm1[ind])
    ws.cell(ind + 1, 2, clm2[ind])
    ws.cell(ind + 1, 3, clm3[ind])

wb1.save('results.xlsx')  # saves virtual workbook to a file


"""
# This is a code for csv type files
filename = 'data.csv'
results_f = 'results.csv'
with open(filename) as f:
    reader = csv.reader(f)
    header_row = next(reader)
    main_l = []
    for row in reader:
        temp_l = []
        for ind, head in enumerate(header_row):
            temp_l.append(row[ind])
        main_l.append(temp_l)
    f.close()

print(len(main_l))
new_l = []
clm1, clm2, clm3 = [], [], []
for i in range(len(header_row)-1):
    for row in main_l:
        clm1.append(row[0])
        clm2.append(header_row[i+1])
        clm3.append(row[i+1])

with open(results_f, mode='w') as results:
    writer = csv.writer(results, delimiter=',')
    for ind in range(len(clm1)):
        writer.writerow([clm1[ind], clm2[ind], clm3[ind]])
    results.close()
"""